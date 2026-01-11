"""
Script to read Excel file and fill 15-minute OHLC data
Picks dates from Column B and fills First 15 Min OHLC in columns H-K
"""

import openpyxl
from datetime import datetime
from growwapi import GrowwAPI
import time

# Configuration
EXCEL_FILE_PATH = '/Users/amankumarayush/Downloads/OHLC.xlsx'
OUTPUT_FILE_PATH = '/Users/amankumarayush/Downloads/OHLC_filled.xlsx'



def initialize_groww_api():
    """Initialize and authenticate with Groww API"""
    try:
        access_token = "eyJraWQiOiJaTUtjVXciLCJhbGciOiJFUzI1NiJ9.eyJleHAiOjE3NjQzNzYyMDAsImlhdCI6MTc2NDM0NTM4NiwibmJmIjoxNzY0MzQ1Mzg2LCJzdWIiOiJ7XCJ0b2tlblJlZklkXCI6XCI3ZTJlOTM4NC05Nzk4LTQxNjgtOGQ2NS1iZjEzMTZhYzhmZjNcIixcInZlbmRvckludGVncmF0aW9uS2V5XCI6XCJlMzFmZjIzYjA4NmI0MDZjODg3NGIyZjZkODQ5NTMxM1wiLFwidXNlckFjY291bnRJZFwiOlwiNTY1NmU1ODEtZmI4NC00OGYyLTk5NjctYWExODQ2MmQ4ZTdiXCIsXCJkZXZpY2VJZFwiOlwiY2RlZmFlODMtNjQ2Yy01MjU1LThmMmUtNzYzY2RkN2Y2ZWZhXCIsXCJzZXNzaW9uSWRcIjpcImU1ZTFiOTNiLThiNjktNGU5Yy1iNDA3LTNmODhjNDVlMTIzNlwiLFwiYWRkaXRpb25hbERhdGFcIjpcIno1NC9NZzltdjE2WXdmb0gvS0EwYkVzQmpMRHhxV29paXlMK0ZHb0N6TU5STkczdTlLa2pWZDNoWjU1ZStNZERhWXBOVi9UOUxIRmtQejFFQisybTdRPT1cIixcInJvbGVcIjpcIm9yZGVyLWJhc2ljLGxpdmVfZGF0YS1iYXNpYyxub25fdHJhZGluZy1iYXNpYyxvcmRlcl9yZWFkX29ubHktYmFzaWMsYmFja190ZXN0XCIsXCJzb3VyY2VJcEFkZHJlc3NcIjpcIjJhMDE6Yjc0MDoxM2M2Ojo3MCwxNzIuNjguMjM5LjE5MCwzNS4yNDEuMjMuMTIzXCIsXCJ0d29GYUV4cGlyeVRzXCI6MTc2NDM3NjIwMDAwMH0iLCJpc3MiOiJhcGV4LWF1dGgtcHJvZC1hcHAifQ.ihKbfd8L-1kS5PiX4WWMAiSkzVkow863lBVr9mvMa5KhluVZzwMyhsfVunsSiguz-n0SJXNXO2cn5DIemWPvXg"
        groww = GrowwAPI(access_token)
        print("✓ Successfully authenticated with Groww API")
        return groww
    except Exception as e:
        print(f"✗ Failed to authenticate with Groww API: {e}")
        return None


def get_15min_ohlc_data(groww, trading_symbol, date_obj):
    """
    Fetch first 15-minute OHLC candle data for a given symbol and date
    
    Args:
        groww: GrowwAPI instance
        trading_symbol: Trading symbol (e.g., "NIFTY")
        date_obj: datetime object
    
    Returns:
        dict: OHLC data with keys 'open', 'high', 'low', 'close'
    """
    try:
        # Format times for the first 15 minutes of trading (9:15 AM to 9:30 AM IST)
        start_time = f"{date_obj.strftime('%Y-%m-%d')} 09:15:00"
        end_time = f"{date_obj.strftime('%Y-%m-%d')} 09:30:00"
        
        # Fetch historical candle data for 15-minute interval
        historical_response = groww.get_historical_candle_data(
            trading_symbol=trading_symbol,
            exchange=groww.EXCHANGE_NSE,
            segment=groww.SEGMENT_CASH,
            start_time=start_time,
            end_time=end_time,
            interval_in_minutes=15  # 15-minute interval
        )
        
        if historical_response and 'candles' in historical_response and len(historical_response['candles']) > 0:
            # Get the first candle (first 15 minutes)
            candle = historical_response['candles'][0]
            # Candle format: [timestamp, open, high, low, close, volume]
            return {
                'open': candle[1],
                'high': candle[2],
                'low': candle[3],
                'close': candle[4]
            }
        else:
            print(f"  ⚠ No candle data found for {trading_symbol} on {date_obj.strftime('%Y-%m-%d')}")
            return None
            
    except Exception as e:
        print(f"  ✗ Error fetching historical data: {e}")
        return None


def process_excel_file(excel_path, output_path):
    """
    Read Excel file, fetch OHLC data for dates in Column B, and fill columns H-K
    
    Structure:
    - Column A: Index Name
    - Column B: Date
    - Columns C-F: Daily OHLC (already filled)
    - Column G: Empty
    - Column H: First 15 Min - Open (to be filled)
    - Column I: First 15 Min - High (to be filled)
    - Column J: First 15 Min - Low (to be filled)
    - Column K: First 15 Min - Close (to be filled)
    """
    try:
        # Load workbook
        print(f"Reading Excel file: {excel_path}")
        wb = openpyxl.load_workbook(excel_path)
        
        print(f"✓ Excel file loaded successfully")
        print(f"Sheets found: {wb.sheetnames}")
        
        # Initialize Groww API
        groww = initialize_groww_api()
        if not groww:
            print("Cannot proceed without Groww API authentication")
            return
        
        # Process each sheet
        for sheet_name in wb.sheetnames:
            print(f"\n{'='*80}")
            print(f"Processing Sheet: {sheet_name}")
            print(f"{'='*80}")
            
            ws = wb[sheet_name]
            
            # Row 2 contains headers, data starts from row 3
            total_rows = ws.max_row
            data_rows = total_rows - 2  # Excluding header rows
            
            print(f"Total data rows: {data_rows}")
            
            filled_count = 0
            skipped_count = 0
            error_count = 0
            
            # Process each data row (starting from row 3)
            for row_idx in range(3, total_rows + 1):
                # Get date from Column B (column 2)
                date_cell = ws.cell(row=row_idx, column=2)
                date_value = date_cell.value
                
                if not date_value or not isinstance(date_value, datetime):
                    print(f"Row {row_idx}: Skipping - Invalid or missing date")
                    skipped_count += 1
                    continue
                
                # Check if 15-min OHLC is already filled (Column H)
                existing_open = ws.cell(row=row_idx, column=8).value
                if existing_open is not None:
                    print(f"Row {row_idx}: Skipping - Already has 15-min data for {date_value.strftime('%Y-%m-%d')}")
                    skipped_count += 1
                    continue
                
                date_str = date_value.strftime('%Y-%m-%d')
                print(f"Row {row_idx}: Fetching 15-min OHLC for {date_str}...", end=" ")
                
                # Fetch OHLC data
                trading_symbol = "NIFTY"  # NIFTY 50 index
                ohlc_data = get_15min_ohlc_data(groww, trading_symbol, date_value)
                
                if ohlc_data and all(v is not None for v in ohlc_data.values()):
                    # Fill columns H, I, J, K with 15-min OHLC
                    ws.cell(row=row_idx, column=8, value=ohlc_data['open'])   # Column H - Open
                    ws.cell(row=row_idx, column=9, value=ohlc_data['high'])   # Column I - High
                    ws.cell(row=row_idx, column=10, value=ohlc_data['low'])   # Column J - Low
                    ws.cell(row=row_idx, column=11, value=ohlc_data['close']) # Column K - Close
                    
                    print(f"✓ O={ohlc_data['open']:.2f}, H={ohlc_data['high']:.2f}, "
                          f"L={ohlc_data['low']:.2f}, C={ohlc_data['close']:.2f}")
                    filled_count += 1
                else:
                    print(f"✗ Failed to fetch complete data")
                    error_count += 1
                
                # Rate limiting - sleep to avoid hitting API limits
                time.sleep(0.5)  # 500ms delay between requests
            
            # Sheet summary
            print(f"\n--- Sheet '{sheet_name}' Summary ---")
            print(f"Total rows processed: {data_rows}")
            print(f"Successfully filled: {filled_count}")
            print(f"Skipped (already filled or invalid): {skipped_count}")
            print(f"Errors: {error_count}")
        
        # Save to output file
        print(f"\n{'='*80}")
        print(f"Saving results to: {output_path}")
        wb.save(output_path)
        print(f"✓ Successfully saved filled OHLC data")
        
        wb.close()
        
    except FileNotFoundError:
        print(f"✗ Error: File not found at {excel_path}")
    except Exception as e:
        print(f"✗ Error processing Excel file: {e}")
        import traceback
        traceback.print_exc()


def main():
    """Main function"""
    print("=" * 80)
    print("15-Minute OHLC Data Filler")
    print("Fills First 15 Min OHLC data in columns H-K")
    print("=" * 80)
    
    # Process the Excel file
    process_excel_file(EXCEL_FILE_PATH, OUTPUT_FILE_PATH)
    
    print("\n" + "=" * 80)
    print("Process completed!")
    print("=" * 80)


if __name__ == "__main__":
    main()
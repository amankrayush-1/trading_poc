import openpyxl

wb = openpyxl.load_workbook('/Users/amankumarayush/Downloads/OHLC_filled.xlsx')
ws = wb['2025']

print("Successfully filled rows in 2025 sheet:")
print("="*80)

for i in range(3, ws.max_row + 1):
    open_val = ws.cell(i, 8).value
    if open_val is not None:
        date_val = ws.cell(i, 2).value
        high_val = ws.cell(i, 9).value
        low_val = ws.cell(i, 10).value
        close_val = ws.cell(i, 11).value
        print(f"Row {i}: {date_val.strftime('%Y-%m-%d')} | O={open_val:.2f} H={high_val:.2f} L={low_val:.2f} C={close_val:.2f}")

wb.close()
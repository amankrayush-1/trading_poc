"""
Script to inspect the Excel file structure
"""
import openpyxl

EXCEL_FILE_PATH = '/Users/amankumarayush/Downloads/OHLC.xlsx'

try:
    # Load workbook
    wb = openpyxl.load_workbook(EXCEL_FILE_PATH)
    
    print("=" * 80)
    print(f"Excel File: {EXCEL_FILE_PATH}")
    print("=" * 80)
    
    print(f"\nNumber of sheets: {len(wb.sheetnames)}")
    print(f"Sheet names: {wb.sheetnames}")
    
    # Inspect each sheet
    for sheet_name in wb.sheetnames:
        print("\n" + "=" * 80)
        print(f"SHEET: {sheet_name}")
        print("=" * 80)
        
        ws = wb[sheet_name]
        
        # Get dimensions
        print(f"Max row: {ws.max_row}")
        print(f"Max column: {ws.max_column}")
        
        # Get headers (first row)
        print("\nHeaders (Row 1):")
        headers = []
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=1, column=col).value
            headers.append(cell_value)
            print(f"  Column {col} ({chr(64+col)}): {cell_value}")
        
        # Show first 5 data rows
        print(f"\nFirst 5 data rows:")
        for row in range(2, min(7, ws.max_row + 1)):
            row_data = []
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=row, column=col).value
                row_data.append(cell_value)
            print(f"  Row {row}: {row_data}")
    
    wb.close()
    print("\n" + "=" * 80)
    
except FileNotFoundError:
    print(f"Error: File not found at {EXCEL_FILE_PATH}")
except Exception as e:
    print(f"Error: {e}")
    import traceback
    traceback.print_exc()
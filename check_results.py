import openpyxl

wb = openpyxl.load_workbook('/Users/amankumarayush/Downloads/OHLC_filled.xlsx')

for sheet_name in ['2024', '2025']:
    ws = wb[sheet_name]
    filled = 0
    total = 0
    
    print(f"\n{'='*60}")
    print(f"Sheet: {sheet_name}")
    print(f"{'='*60}")
    
    for i in range(3, min(10, ws.max_row + 1)):  # Check first 7 rows
        total += 1
        date_val = ws.cell(i, 2).value
        open_val = ws.cell(i, 8).value
        high_val = ws.cell(i, 9).value
        low_val = ws.cell(i, 10).value
        close_val = ws.cell(i, 11).value
        
        if open_val is not None:
            filled += 1
            print(f"Row {i}: {date_val} | O={open_val:.2f} H={high_val:.2f} L={low_val:.2f} C={close_val:.2f}")
        else:
            print(f"Row {i}: {date_val} | No data")
    
    # Count total filled
    total_filled = 0
    for i in range(3, ws.max_row + 1):
        if ws.cell(i, 8).value is not None:
            total_filled += 1
    
    print(f"\nTotal rows: {ws.max_row - 2}")
    print(f"Filled rows: {total_filled}")
    print(f"Empty rows: {ws.max_row - 2 - total_filled}")

wb.close()